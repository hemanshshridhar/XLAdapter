import os
from openpyxl import load_workbook
from shutil import rmtree

from llmadapter.analyzer.agent import AnalyzerAgent
from llmadapter.utils.blob_storage_util import BlobStorageUtil
from llmadapter.utils.notifier import send_setup_notification
from llmadapter.utils.util import delete_folder
from llmadapter.tools.sheetencoder import SheetEncoder
from .constants import Status

class LLMadapt:
    def __init__(self):
        self.analyzer = AnalyzerAgent()
        self.sheet_encoder = SheetEncoder()

    def get_sheetnames(self, excel_path):
        """
        Displays available sheets and returns a user-selected list by index.
        """
        wb = load_workbook(excel_path, read_only=True, keep_vba=True)
        all_sheets = wb.sheetnames

        print("\n Available sheets in the Excel file:")
        for idx, name in enumerate(all_sheets, start=1):
            print(f"{idx}. {name}")

        selected_indices = input("\n Enter sheet numbers to select (comma-separated): ")

        try:
            selected_indices = [int(i.strip()) for i in selected_indices.split(",") if i.strip().isdigit()]
        except ValueError:
            print(" Invalid input format. Please enter numbers only.")
            return []

        selected_sheets = [all_sheets[i - 1] for i in selected_indices if 1 <= i <= len(all_sheets)]

        if not selected_sheets:
            print(" No valid selections made. Exiting.")
            exit()

        print(f" Selected sheets: {selected_sheets}")
        return selected_sheets

    def model_setup(self, input_data_model_setup):
        model_id            = input_data_model_setup.get("model_id")
        base_excel_path     = input_data_model_setup.get("model_excel_path")
        country_excel_path  = input_data_model_setup.get("sheets_path")
        output_excel_path   = input_data_model_setup.get("output_excel_path")
        output_log_path     = input_data_model_setup.get("output_log_path")
        # Prompt user to select sheets interactively
        sheetnames = self.get_sheetnames(base_excel_path)

        print(f"[INFO] Starting model setup for: {model_id}")

        try:
            output_path = self.compare_and_generate_updated_xlsm(
                base_file_path    = base_excel_path,
                country_file_path = country_excel_path,
                output_file_path  = output_excel_path,
                output_log_path   = output_log_path,
                sheetnames        = sheetnames
            )
            print(f"[SUCCESS] Model setup completed for: {model_id}")
        except Exception as e:
            print(f"[FAILED] Model setup failed for: {model_id} — {e}")
            raise
        # finally:
        #     temp_dir = os.path.dirname(base_excel_path)
        #     if os.path.exists(temp_dir):
        #         rmtree(temp_dir, ignore_errors=True)

        return output_path

    def compare_and_generate_updated_xlsm(
        self,
        base_file_path: str,
        country_file_path: str,
        output_file_path: str,
        sheetnames: list,
        output_log_path: str
    ):
        """
        Compare two .xlsm files via LLM, generate a result dictionary, and write to a new .xlsm file.
        """
        # Step 1: Convert Excel sheets to dictionaries
        model_dict   = self.sheet_encoder.encode_model(base_file_path, sheetnames)
        country_dict = self.sheet_encoder.encode_sheet(country_file_path)
        output = self.analyzer.process(model_dict, country_dict, sheetnames)
        print(output)
        output_dict = output['output_dict']
        log_table = output['log_table']
        print(log_table)

        # Step 2: Write the result to new Excel file
        self.sheet_encoder.write_to_sheet(
            output_dict    = output_dict,
            template_path  = base_file_path,
            output_path    = output_file_path
        )
        self.sheet_encoder.write_log_table(
            log_table = log_table,
            output_path = output_log_path
        )

        return output_file_path
