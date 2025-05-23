import pandas as pd
from openpyxl import load_workbook
from typing import Dict, List, Any
import json

class SheetEncoder:
    def __init__(self):
        pass

    def encode_model(self, filename: str, sheet_names: List[str] = None) -> str:
        wb_formula = load_workbook(filename, keep_vba=True)
        wb_value = load_workbook(filename, data_only=True)

        if sheet_names is None:
            sheet_names = wb_formula.sheetnames

        excel_data = {
            "workbook_name": filename,
            "sheets": []
        }

        for sheet_name in sheet_names:
            if sheet_name not in wb_formula.sheetnames:
                print(f"Warning: Sheet '{sheet_name}' not found in workbook.")
                continue

            sheet_formula = wb_formula[sheet_name]
            sheet_value = wb_value[sheet_name]

            sheet_data = {
                "name": sheet_name,
                "cell_data": []
            }

            for row in sheet_formula.iter_rows():
                for cell_formula in row:
                    coord = cell_formula.coordinate
                    cell_value = sheet_value[coord]

                    if cell_formula.value is None and cell_value.value is None:
                        continue

                    cell_info = {
                        "address": coord,
                        "value": cell_value.value,
                        "formula": str(cell_formula.value) if cell_formula.data_type == 'f' else None
                    }

                    sheet_data["cell_data"].append(cell_info)

            excel_data["sheets"].append(sheet_data)
            json_data = json.dumps(excel_data, indent=4)

        return excel_data

    def encode_sheet(self, sheet_path: str) -> Dict[str, Dict[str, Any]]:
        wb = load_workbook(sheet_path, data_only=True)
        ws = wb.active  # Use the first sheet

        raw_data = []
        for row in ws.iter_rows(values_only=True):
            raw_data.append(row)

        df = pd.DataFrame(raw_data)
        df = df.iloc[2:, [2, 3, 4]]  # Columns C, D, E
        df.columns = ['sheet_name', 'parameter', 'value']
        df['sheet_name'] = df['sheet_name'].fillna(method='ffill')
        df = df.dropna(subset=['parameter'])

        structured_dict = {}
        for _, row in df.iterrows():
            sheet = row['sheet_name']
            param = row['parameter']
            val = row['value']
            structured_dict.setdefault(sheet, {})[param] = val

        return structured_dict

    def write_to_sheet(
        self,
        output_dict: Dict[str, Dict[str, Any]],
        template_path: str,
        output_path: str,
        keep_vba: bool = True
    ):
        wb = load_workbook(template_path, keep_vba=keep_vba)

        for sheet_name, cell_map in output_dict.items():
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found in workbook")

            ws = wb[sheet_name]

            for addr, val in cell_map.items():
                ws[addr].value = val

        wb.save(output_path)

    def write_log_table(
        self,
        log_table: List[Dict[str, Any]],
        output_path: str = "change_log.xlsx",
    ):
        """
        Write the change-log table to an Excel file.

        Parameters
        ----------
        log_table : list[dict]
            Each dict should contain keys like:
            'Sheet Name', 'Cell Field', 'Cell Address',
            'Previous Value', 'New Value'
        output_path : str
            Path where the Excel file will be saved.
        """
        if not log_table or not isinstance(log_table, list):
            raise ValueError("log_table must be a non-empty list of dictionaries.")

        # 🔄 Replace Python None with the literal string "None"
        for row in log_table:
            for key, val in row.items():
                if val is None:
                    row[key] = "None"

        df = pd.DataFrame(log_table)
        df.to_excel(output_path, index=False)
        print(f"✓ Log table written to: {output_path}")
