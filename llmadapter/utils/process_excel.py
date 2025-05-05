from openpyxl import load_workbook
from typing import Dict, Any
from oletools.olevba import VBA_Parser
import json

def process_excel(filename: str) -> Dict[str, Any]:
    wb_formula = load_workbook(filename, keep_vba=True)
    wb_value = load_workbook(filename, data_only=True)

    excel_data = {
        "workbook_name": filename,
        "sheets": []
    }

    for sheet_name in wb_formula.sheetnames:
        sheet_formula = wb_formula[sheet_name]
        sheet_value = wb_value[sheet_name]

        sheet_data = {
            "name": sheet_name,
            "cell_data": []
        }

        for row in sheet_formula.iter_rows():
            for cell_formula in row:
                coord = cell_formula.coordinate
                cell_value = sheet_value[coord]

                if cell_formula.value is None and cell_value.value is None:
                    continue

                cell_info = {
                    "address": coord,
                    "value": cell_value.value,
                    "formula": str(cell_formula.value) if cell_formula.data_type == 'f' else None
                }

                sheet_data["cell_data"].append(cell_info)

        excel_data["sheets"].append(sheet_data)

    return excel_data

def generate_json_from_excel(excel_file_path, json_file_path):
    excel_data = process_excel(filename=excel_file_path)
    json_object = json.dumps(excel_data, indent=4)
    with open(json_file_path, "w") as file:
        file.write(json_object)

    print(f"{excel_file_path} is successfully extracted and saved to {json_file_path}")

def extract_vba_from_excel(excel_file_path):
    vbaparser = VBA_Parser(excel_file_path)

    if not vbaparser.detect_vba_macros():
        print("No VBA macros found in this workbook.")
        return "No VBA macros found in this workbook."

    content = ''
    for (subfilename, stream_path, vba_filename, vba_code) in vbaparser.extract_all_macros():
        # Filter out metadata lines that start with "Attribute"
        clean_code = "\n".join(
            line for line in vba_code.splitlines() if not line.strip().startswith("Attribute")
        )

        if clean_code:
            content += f"Module: {vba_filename} (from stream: {stream_path})\n"
            content += clean_code + "\n"
            content += "=" * 80 + "\n"
    return content
